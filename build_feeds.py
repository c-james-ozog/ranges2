"""
ranges2 feed builder.

Fetches daily OHLC data from Yahoo Finance for all configured contracts,
applies manual overrides, computes targets/achievements/volatility metrics,
and writes JSON feeds to the feeds/ directory.

Usage:
    python build_feeds.py

Output:
    feeds/history/<SYMBOL>.json   — per-contract history
    feeds/overview-by-date.json   — all contracts indexed by date
    feeds/meta.json               — build metadata
    feeds/errors.json             — any per-contract errors
"""

from __future__ import annotations

import json
import logging
import time
from concurrent.futures import Future, ThreadPoolExecutor
from datetime import date, datetime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

from src.config import (
    CME_HOLIDAYS,
    CONTRACT_BY_SYMBOL,
    CONTRACTS,
    DAILY_TARGET_LOOKBACK,
    FETCH_DELAY,
    FETCH_WORKERS,
    HOME_ORDER,
    HV_ANNUALIZATION_FACTOR,
    HV_TARGET_MULTIPLIER,
    PRICE_DIVISOR,
    TICK_SIZES,
    WEEKLY_TARGET_LOOKBACK,
    Contract,
    active_symbols_for_date,
)
from src.scraper import RawRow, fetch_yahoo_history, format_tick, round_to_tick

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)-8s %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

FEEDS_DIR    = Path("feeds")
HISTORY_DIR  = FEEDS_DIR / "history"

IMPLIED_VOL_FILE    = Path("implied_vol.json")
PRICE_OVERRIDE_FILE = Path("price_overrides.json")
EXCEL_OVERRIDE_FILE = Path("implied_vol_input.xlsx")

# ---------------------------------------------------------------------------
# Chicago timezone & date helpers
# ---------------------------------------------------------------------------

CT = ZoneInfo("America/Chicago")


def ts_to_ct_date(ts: int) -> str:
    """
    Convert a Yahoo Finance Unix timestamp to a Chicago-time trade date.
    Yahoo stores timestamps at midnight UTC for the *previous* calendar day,
    so adding one day corrects to the actual trade date.
    """
    return (datetime.fromtimestamp(ts, tz=CT) + timedelta(days=1)).strftime("%Y-%m-%d")


def is_trading_day(date_str: str) -> bool:
    """Return True if date_str is a weekday that is not a CME holiday."""
    try:
        d = date.fromisoformat(date_str)
    except ValueError:
        return False
    return d.weekday() < 5 and date_str not in CME_HOLIDAYS


def week_monday(date_str: str) -> str:
    """Return the ISO date of the Monday that starts the week containing date_str."""
    d = date.fromisoformat(date_str)
    return (d - timedelta(days=d.weekday())).isoformat()


def is_last_trading_day_of_week(date_str: str) -> bool:
    """True if date_str is the last trading day (Friday or earlier if holiday) of its week."""
    d = date.fromisoformat(date_str)
    friday = d + timedelta(days=4 - d.weekday())
    candidate = friday
    for _ in range(5):
        iso = candidate.isoformat()
        if candidate.weekday() < 5 and iso not in CME_HOLIDAYS:
            return date_str == iso
        candidate -= timedelta(days=1)
    return False


def is_first_trading_day_of_week(date_str: str) -> bool:
    """True if date_str is the first trading day (Monday or later if holiday) of its week."""
    d = date.fromisoformat(date_str)
    monday = d - timedelta(days=d.weekday())
    candidate = monday
    for _ in range(5):
        iso = candidate.isoformat()
        if candidate.weekday() < 5 and iso not in CME_HOLIDAYS:
            return date_str == iso
        candidate += timedelta(days=1)
    return False


# ---------------------------------------------------------------------------
# Override loaders
# ---------------------------------------------------------------------------

def load_json_file(path: Path) -> dict:
    """Load a JSON file, returning an empty dict on missing file or parse error."""
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError) as e:
        log.warning("Could not load %s: %s", path, e)
        return {}


def load_implied_vol() -> dict[str, dict[str, float]]:
    """
    Load implied volatility data from implied_vol.json.
    Format: { "YYYY-MM-DD": { "SYMBOL": float_percent } }
    """
    data = load_json_file(IMPLIED_VOL_FILE)
    if data:
        log.info("Loaded implied vol for %d dates", len(data))
    return data


def load_price_overrides() -> dict[str, dict[str, dict[str, float]]]:
    """
    Load manual price corrections from price_overrides.json.
    Format: { "YYYY-MM-DD": { "SYMBOL": { "high": float, "low": float } } }
    """
    data = load_json_file(PRICE_OVERRIDE_FILE)
    if data:
        log.info("Loaded price overrides for %d dates", len(data))
    return data


def load_rice_overrides() -> dict[str, dict[str, float]]:
    """
    Load Rice (ZRN26) manual OHLC data from the Excel workbook.
    Yahoo Finance data for Rice is unreliable; this provides correct values.
    Format: { "YYYY-MM-DD": { "high": float, "low": float, "close": float } }
    """
    if not EXCEL_OVERRIDE_FILE.exists():
        return {}
    try:
        from openpyxl import load_workbook  # lazy import — only needed if file exists
        wb = load_workbook(EXCEL_OVERRIDE_FILE, data_only=True)
        sheet_name = "Rice Override (ZRN26)"
        if sheet_name not in wb.sheetnames:
            return {}
        ws = wb[sheet_name]
        result: dict[str, dict[str, float]] = {}
        for date_val, high, low, close, *_ in ws.iter_rows(min_row=3, values_only=True):
            if date_val is None or high is None or low is None:
                continue
            date_str = (
                date_val.strftime("%Y-%m-%d")
                if hasattr(date_val, "strftime")
                else datetime.strptime(str(date_val), "%m/%d/%y").strftime("%Y-%m-%d")
            )
            result[date_str] = {
                "high": float(high),
                "low": float(low),
                "close": float(close) if close is not None else 0.0,
            }
        if result:
            log.info("Loaded Rice overrides for %d dates", len(result))
        return result
    except Exception as e:
        log.warning("Could not load Rice overrides from Excel: %s", e)
        return {}


# ---------------------------------------------------------------------------
# Row pre-processing
# ---------------------------------------------------------------------------

def preprocess_rows(
    raw_rows: list[RawRow],
    symbol: str,
    today: str,
    price_overrides: dict,
    rice_overrides: dict,
) -> list[RawRow]:
    """
    Apply all corrections to raw Yahoo rows and filter to valid trading days <= today.

    Processing order:
        1. Apply price_overrides (corrects bad Yahoo highs/lows for any contract)
        2. Apply rice_overrides (replaces all Yahoo data for ZRN26)
        3. Filter: keep only valid trading days on or before today
    """
    result: list[RawRow] = []
    for row in raw_rows:
        trade_date = ts_to_ct_date(row["timestamp"])

        # Skip future dates and non-trading days
        if trade_date > today or not is_trading_day(trade_date):
            continue

        # Apply generic price overrides
        if day := price_overrides.get(trade_date, {}).get(symbol):
            row = dict(row)  # avoid mutating original
            row["high"]  = float(day.get("high",  row["high"]))
            row["low"]   = float(day.get("low",   row["low"]))
            row["close"] = float(day.get("close", row["close"]))

        # Apply Rice-specific overrides
        if symbol == "ZRN26" and (override := rice_overrides.get(trade_date)):
            row = dict(row)
            row["high"]  = override["high"]
            row["low"]   = override["low"]
            row["close"] = override["close"] or row["close"]

        result.append(row)

    return result


# ---------------------------------------------------------------------------
# Tick-level computations
# ---------------------------------------------------------------------------

def daily_range(row: RawRow, tick: float) -> float:
    return round_to_tick(round_to_tick(row["high"], tick) - round_to_tick(row["low"], tick), tick)


def pct(numerator: float | None, denominator: float | None) -> str:
    """Format numerator/denominator as a percentage string, e.g. '95.5%'."""
    if not numerator or not denominator:
        return ""
    return f"{round((numerator / denominator) * 100, 1)}%"


def historic_vol(rows: list[RawRow], i: int, tick: float) -> str:
    """
    Historic Vol = (avg_3day_range * HV_TARGET_MULT) / (close / PRICE_DIVISOR) * HV_ANNUALIZATION

    Uses a rolling window of DAILY_TARGET_LOOKBACK days starting at index i.
    Returns '' if insufficient data.
    """
    window = rows[i : i + DAILY_TARGET_LOOKBACK]
    if len(window) < DAILY_TARGET_LOOKBACK or not window[0].get("close"):
        return ""
    avg_range = sum(daily_range(r, tick) for r in window) / DAILY_TARGET_LOOKBACK
    close_one_pct = window[0]["close"] / PRICE_DIVISOR
    hv = (avg_range * HV_TARGET_MULTIPLIER / close_one_pct) * HV_ANNUALIZATION_FACTOR
    return f"{round(hv, 1)}%"


def full_achievement_and_target(
    rows: list[RawRow], i: int, tick: float
) -> tuple[float | None, float | None]:
    """
    Return (avg_range, daily_target) for the window starting at index i.
    avg_range is the average of DAILY_TARGET_LOOKBACK daily ranges.
    daily_target is avg_range * HV_TARGET_MULTIPLIER.
    Returns (None, None) if insufficient data.
    """
    window = rows[i : i + DAILY_TARGET_LOOKBACK]
    if len(window) < DAILY_TARGET_LOOKBACK:
        return None, None
    avg = sum(daily_range(r, tick) for r in window) / DAILY_TARGET_LOOKBACK
    avg_r  = round_to_tick(avg, tick)
    target = round_to_tick(avg_r * HV_TARGET_MULTIPLIER, tick)
    return avg_r, target


# ---------------------------------------------------------------------------
# Weekly computations
# ---------------------------------------------------------------------------

WeeklyData = dict  # date_str -> {weeklyHigh, weeklyLow, weeklyRange}
WeeklyTargets = dict  # date_str -> {weeklyTarget, nextWeeklyTarget}


def compute_weekly_ranges(dated_rows: list[dict], tick: float) -> WeeklyData:
    """
    Compute cumulative weekly high/low/range for each trading day.
    Ranges expand Mon→Fri as new highs/lows are made during the week.
    """
    # Group rows by their week's Monday
    weeks: dict[str, list[dict]] = {}
    for row in dated_rows:
        weeks.setdefault(week_monday(row["date"]), []).append(row)

    result: WeeklyData = {}
    for week_rows in weeks.values():
        week_rows.sort(key=lambda r: r["date"])
        cum_high = cum_low = None
        for row in week_rows:
            h = round_to_tick(row["high"], tick)
            l = round_to_tick(row["low"], tick)
            cum_high = h if cum_high is None else max(cum_high, h)
            cum_low  = l if cum_low  is None else min(cum_low,  l)
            result[row["date"]] = {
                "weeklyHigh":  cum_high,
                "weeklyLow":   cum_low,
                "weeklyRange": round_to_tick(cum_high - cum_low, tick),
            }
    return result


def compute_completed_weeks(dated_rows: list[dict], tick: float) -> list[dict]:
    """
    Return completed weeks (those ending on a valid last trading day), newest first.
    Each entry: {monday, lastDay, high, low, range}
    """
    weeks: dict[str, list[dict]] = {}
    for row in dated_rows:
        weeks.setdefault(week_monday(row["date"]), []).append(row)

    completed = []
    for monday, week_rows in sorted(weeks.items(), reverse=True):
        week_rows.sort(key=lambda r: r["date"])
        last_day = week_rows[-1]["date"]
        if not is_last_trading_day_of_week(last_day):
            continue
        high = round_to_tick(max(r["high"] for r in week_rows), tick)
        low  = round_to_tick(min(r["low"]  for r in week_rows), tick)
        completed.append({
            "monday":  monday,
            "lastDay": last_day,
            "high":    high,
            "low":     low,
            "range":   round_to_tick(high - low, tick),
        })
    return completed


def compute_weekly_targets(dated_rows: list[dict], tick: float) -> WeeklyTargets:
    """
    For each trading day, determine the applicable weekly target and next weekly target.

    Weekly target for week W = avg range of the 3 prior completed weeks.
    nextWeeklyTarget is shown on the last trading day of the week (for Friday display).
    """
    completed = compute_completed_weeks(dated_rows, tick)

    # Map each week's Monday to its target (derived from the 3 prior completed weeks)
    target_by_monday: dict[str, float] = {}
    for i in range(len(completed) - WEEKLY_TARGET_LOOKBACK + 1):
        window = completed[i : i + WEEKLY_TARGET_LOOKBACK]
        avg = sum(w["range"] for w in window) / WEEKLY_TARGET_LOOKBACK
        next_mon = (date.fromisoformat(completed[i]["monday"]) + timedelta(weeks=1)).isoformat()
        target_by_monday[next_mon] = round_to_tick(avg, tick)

    # nextWeeklyTarget: available on the last trading day of a completed week
    next_target_by_last_day: dict[str, float] = {}
    for week in completed:
        next_mon = (date.fromisoformat(week["monday"]) + timedelta(weeks=1)).isoformat()
        if target := target_by_monday.get(next_mon):
            next_target_by_last_day[week["lastDay"]] = target

    return {
        row["date"]: {
            "weeklyTarget":     target_by_monday.get(week_monday(row["date"])),
            "nextWeeklyTarget": next_target_by_last_day.get(row["date"]),
        }
        for row in dated_rows
    }


# ---------------------------------------------------------------------------
# Implied vol trend
# ---------------------------------------------------------------------------

def compute_iv_trends(history_rows: list[dict]) -> list[str]:
    """
    For each row (newest-first) compute the implied vol trend string.

    Format: "up|N" or "down|N" where N is the consecutive-day streak count.
    Returns "" for the oldest data point or when no prior IV exists.

    Processes oldest-to-newest internally for streak counting, then reverses.
    """
    reversed_rows = list(reversed(history_rows))
    trends = [""] * len(reversed_rows)

    for i, row in enumerate(reversed_rows):
        raw = row.get("impliedVol", "")
        if not raw:
            continue
        try:
            iv = float(raw.replace("%", ""))
        except ValueError:
            continue

        # Find the most recent prior row with an IV value
        prev_iv: float | None = None
        for j in range(i - 1, -1, -1):
            prev_raw = reversed_rows[j].get("impliedVol", "")
            if prev_raw:
                try:
                    prev_iv = float(prev_raw.replace("%", ""))
                    break
                except ValueError:
                    continue

        if prev_iv is None:
            continue

        if iv > prev_iv:
            direction = "up"
        elif iv < prev_iv:
            direction = "down"
        else:
            continue  # unchanged — no trend marker

        # Count consecutive days in this direction
        count = 1
        for j in range(i - 1, -1, -1):
            prev_trend = trends[j]
            if not prev_trend or prev_trend.split("|")[0] != direction:
                break
            count += 1

        trends[i] = f"{direction}|{count}"

    return list(reversed(trends))


# ---------------------------------------------------------------------------
# History builder
# ---------------------------------------------------------------------------

def build_history(rows: list[RawRow], contract: Contract, iv_data: dict) -> list[dict]:
    """
    Build the full history row list for a single contract.

    Args:
        rows:     Pre-processed (filtered, overridden) Yahoo rows, newest-first.
        contract: Contract config dict.
        iv_data:  Full implied vol data keyed by date → symbol → float.

    Returns:
        List of history row dicts, newest-first, ready for JSON serialisation.
    """
    tick = TICK_SIZES[contract["commodity"]]
    symbol = contract["base_symbol"]

    # Build lightweight dated rows for weekly computations
    dated = [
        {"date": ts_to_ct_date(r["timestamp"]), "high": r["high"], "low": r["low"]}
        for r in rows
    ]

    weekly_ranges  = compute_weekly_ranges(dated, tick)
    weekly_targets = compute_weekly_targets(dated, tick)

    # ---- First pass: build core row data ----
    history: list[dict] = []
    for i, row in enumerate(rows):
        trade_date = ts_to_ct_date(row["timestamp"])
        d_high = round_to_tick(row["high"], tick)
        d_low  = round_to_tick(row["low"],  tick)
        d_range = round_to_tick(d_high - d_low, tick)

        full_ach, next_target = full_achievement_and_target(rows, i, tick)
        hv = historic_vol(rows, i, tick)

        iv_val = iv_data.get(trade_date, {}).get(symbol)
        iv_str = f"{round(float(iv_val), 1)}%" if iv_val is not None else ""

        wr = weekly_ranges.get(trade_date, {})
        wt = weekly_targets.get(trade_date, {})

        history.append({
            "date":            trade_date,
            "dailyHigh":       format_tick(d_high, tick),
            "dailyLow":        format_tick(d_low,  tick),
            "dailyRange":      format_tick(d_range, tick),
            "fullAchievement": format_tick(full_ach, tick) if full_ach is not None else "",
            "_fullAch":        full_ach,    # temp — used in second pass
            "_nextTarget":     next_target, # temp — used in second pass
            "historicVol":     hv,
            "impliedVol":      iv_str,
            "weeklyHigh":      format_tick(wr["weeklyHigh"],  tick) if wr else "",
            "weeklyLow":       format_tick(wr["weeklyLow"],   tick) if wr else "",
            "weeklyRange":     format_tick(wr["weeklyRange"],  tick) if wr else "",
            "_weeklyTarget":   wt.get("weeklyTarget"),   # temp
            "_nextWeeklyTarget": wt.get("nextWeeklyTarget"), # temp
            "sectionBreak":    False,
        })

    # ---- Second pass: targets, achievements, section breaks ----
    for i, row in enumerate(history):
        prev = history[i + 1] if i + 1 < len(history) else None

        daily_target = prev["_nextTarget"] if prev else None
        d_range_num  = float(row["dailyRange"]) if row["dailyRange"] else None
        w_range_num  = float(row["weeklyRange"]) if row["weeklyRange"] else None
        weekly_target = row["_weeklyTarget"]
        next_weekly   = row["_nextWeeklyTarget"]

        row["dailyTarget"]      = format_tick(daily_target, tick) if daily_target else ""
        row["nextDailyTarget"]  = format_tick(row["_nextTarget"], tick) if row["_nextTarget"] else ""
        row["dailyAchievement"] = pct(d_range_num, daily_target)
        row["weeklyTarget"]     = format_tick(weekly_target, tick) if weekly_target else ""
        row["nextWeeklyTarget"] = format_tick(next_weekly, tick) if next_weekly else ""
        row["weeklyAchievement"]= pct(w_range_num, weekly_target)
        row["sectionBreak"]     = is_first_trading_day_of_week(row["date"])

        # Remove temp keys
        del row["_fullAch"], row["_nextTarget"], row["_weeklyTarget"], row["_nextWeeklyTarget"]

    # ---- Third pass: implied vol trends ----
    trends = compute_iv_trends(history)
    for row, trend in zip(history, trends):
        row["impliedVolTrend"] = trend

    return history


# ---------------------------------------------------------------------------
# Overview row builder
# ---------------------------------------------------------------------------

def to_overview_row(history_row: dict, contract: Contract) -> dict:
    """Flatten a history row + contract metadata into an overview row."""
    return {
        "date":            history_row["date"],
        "symbol":          contract["base_symbol"],
        "commodity":       contract["commodity"],
        "month":           contract["month"],
        "dailyTarget":     history_row["dailyTarget"],
        "dailyRange":      history_row["dailyRange"],
        "dailyHigh":       history_row["dailyHigh"],
        "dailyLow":        history_row["dailyLow"],
        "dailyAchievement":history_row["dailyAchievement"],
        "fullAchievement": history_row["fullAchievement"],
        "nextDailyTarget": history_row["nextDailyTarget"],
        "historicVol":     history_row["historicVol"],
        "impliedVol":      history_row["impliedVol"],
        "impliedVolTrend": history_row["impliedVolTrend"],
        "weeklyRange":     history_row["weeklyRange"],
        "weeklyHigh":      history_row["weeklyHigh"],
        "weeklyLow":       history_row["weeklyLow"],
        "weeklyAchievement":history_row["weeklyAchievement"],
        "weeklyTarget":    history_row["weeklyTarget"],
        "nextWeeklyTarget":history_row["nextWeeklyTarget"],
    }


# ---------------------------------------------------------------------------
# Contract processor  (runs in thread pool)
# ---------------------------------------------------------------------------

def process_contract(
    contract: Contract,
    today: str,
    price_overrides: dict,
    rice_overrides: dict,
    iv_data: dict,
) -> tuple[str, list[dict] | Exception]:
    """
    Fetch and process a single contract. Returns (base_symbol, history_rows | Exception).
    Designed to run in a thread pool — all inputs are read-only.
    """
    symbol = contract["base_symbol"]
    try:
        time.sleep(FETCH_DELAY)
        raw_rows = fetch_yahoo_history(contract["symbol"])
        clean_rows = preprocess_rows(raw_rows, symbol, today, price_overrides, rice_overrides)
        history = build_history(clean_rows, contract, iv_data)
        log.info("OK  %s  (%d rows)", symbol, len(history))
        return symbol, history
    except Exception as exc:
        log.error("ERR %s: %s", symbol, exc)
        return symbol, exc


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    FEEDS_DIR.mkdir(exist_ok=True)
    HISTORY_DIR.mkdir(exist_ok=True)

    today    = datetime.now(CT).strftime("%Y-%m-%d")
    now_ct   = datetime.now(CT).isoformat()

    log.info("Building feeds for %s (today = %s)", now_ct[:10], today)

    # Load all override/supplement data upfront
    iv_data         = load_implied_vol()
    price_overrides = load_price_overrides()
    rice_overrides  = load_rice_overrides()

    # ---- Parallel fetch & process ----
    results: dict[str, list[dict]] = {}
    errors:  list[dict] = []

    with ThreadPoolExecutor(max_workers=FETCH_WORKERS) as pool:
        future_map: dict[Future, Contract] = {
            pool.submit(
                process_contract,
                contract, today, price_overrides, rice_overrides, iv_data,
            ): contract
            for contract in CONTRACTS
        }
        # Collect results preserving HOME_ORDER by iterating in submission order
        for future in future_map:
            symbol, result = future.result()
            if isinstance(result, Exception):
                errors.append({"symbol": symbol, "error": str(result)})
            else:
                results[symbol] = result

    # ---- Write per-contract history feeds ----
    history_index: list[str] = []
    all_overview_rows: list[dict] = []

    for contract in CONTRACTS:
        symbol = contract["base_symbol"]
        if symbol not in results:
            continue

        history_rows = results[symbol]
        payload = {
            "symbol":    symbol,
            "commodity": contract["commodity"],
            "month":     contract["month"],
            "updatedAt": now_ct,
            "rows":      history_rows,
        }
        (HISTORY_DIR / f"{symbol}.json").write_text(
            json.dumps(payload, indent=2), encoding="utf-8"
        )

        all_overview_rows.extend(
            to_overview_row(row, contract) for row in history_rows
        )
        history_index.append(symbol)

    # ---- Write overview-by-date feed (date-aware contract selection) ----
    # For each date, include the preferred contract per commodity.
    # Contracts with always_show=True (e.g. ZCZ26, ZSX26) are always included.
    # If the preferred contract has no data (expired), fall back to any contract
    # for that commodity that does have data.

    # Symbols that always appear regardless of roll logic
    always_show_syms = {c["base_symbol"] for c in CONTRACTS if c.get("always_show")}

    # Index all rows by (date, symbol)
    rows_by_date_sym: dict[str, dict[str, dict]] = {}
    for row in all_overview_rows:
        rows_by_date_sym.setdefault(row["date"], {})[row["symbol"]] = row

    overview_by_date: dict[str, list[dict]] = {}
    for date_str, sym_map in rows_by_date_sym.items():
        preferred = active_symbols_for_date(date_str)
        date_rows = []
        added_syms: set[str] = set()
        seen_commodities: list[str] = []

        for sym in preferred:
            contract = CONTRACT_BY_SYMBOL.get(sym)
            if not contract:
                continue
            commodity = contract["commodity"]
            is_always = sym in always_show_syms

            # always_show contracts bypass the one-per-commodity rule
            if not is_always and commodity in seen_commodities:
                continue

            if sym in sym_map:
                date_rows.append(sym_map[sym])
                added_syms.add(sym)
                if not is_always:
                    seen_commodities.append(commodity)
            elif not is_always:
                # Preferred contract has no data — fall back
                for c in CONTRACTS:
                    if c["commodity"] == commodity and c["base_symbol"] in sym_map:
                        date_rows.append(sym_map[c["base_symbol"]])
                        added_syms.add(c["base_symbol"])
                        seen_commodities.append(commodity)
                        break

        # Ensure always_show contracts are included even if not in preferred list
        for sym in always_show_syms:
            if sym not in added_syms and sym in sym_map:
                date_rows.append(sym_map[sym])

        # Sort by HOME_ORDER
        order_map = {sym: i for i, sym in enumerate(HOME_ORDER)}
        date_rows.sort(key=lambda r: order_map.get(r["symbol"], 9999))
        if date_rows:
            overview_by_date[date_str] = date_rows

    (FEEDS_DIR / "overview-by-date.json").write_text(
        json.dumps(overview_by_date, indent=2), encoding="utf-8"
    )

    # ---- Write ancillary feeds ----
    (FEEDS_DIR / "errors.json").write_text(
        json.dumps(errors, indent=2), encoding="utf-8"
    )
    (HISTORY_DIR / "index.json").write_text(
        json.dumps({"contracts": history_index}, indent=2), encoding="utf-8"
    )
    (FEEDS_DIR / "meta.json").write_text(
        json.dumps({
            "builtAt":      now_ct,
            "status":       "ok" if not errors else "partial",
            "successCount": len(results),
            "errorCount":   len(errors),
            "version":      "v4.0",
        }, indent=2),
        encoding="utf-8",
    )

    log.info(
        "Done — %d/%d contracts (%d errors)",
        len(results), len(CONTRACTS), len(errors),
    )
    if errors:
        for e in errors:
            log.warning("  Failed: %s — %s", e["symbol"], e["error"])


if __name__ == "__main__":
    main()
